from openpyxl import load_workbook
from dataclasses import dataclass
import win32com.client       # Need pywin32 from pip
from PIL import ImageGrab    # Need PIL as well
import os
import xlwings as xw
from xlwings.constants import CopyPictureFormat, PictureAppearance


@dataclass
class CustomerSheet:
    s_no: int
    list_no: int
    sublist_no: int
    desc: str
    model: str
    dimensions: str
    remarks: str
    unit: str
    unit_price: int
    qty: int
    total_amt: int
    discount_amt: int
    token_cus: int
    token_hstc: int
    token_date: int
    cbm: int
    ctns: int
    gross_wt: int
    net_wt: int	
    photos: str



workbook = load_workbook(filename=r"C:\Users\bharath878\Desktop\Work Files\as\Freelancing\HSTC\Excel Files\VendorSheet.xlsx", data_only=True)
print(workbook.sheetnames)
print('\n')
sheet = workbook.active

customer_sheet = []

print(sheet['M3'].value)

img = xw.Range('M3:M3').api.CopyPicture(PictureAppearance.xlScreen,
                                  CopyPictureFormat.xlPicture)
if img:
    print("image copied")
    print(img)
    # image = ImageGrab.grabclipboard()
    img.save(r"C:\Users\bharath878\Desktop\Work Files\WDF\hstc\hstc\Frontend\python",'jpeg')

for row in sheet.iter_rows(min_row=3, max_row=4, values_only=True):
    # row.CopyPicture()

    # image = ImageGrab.grabclipboard()
    # image.save(r"C:\Users\bharath878\Desktop\Work Files\WDF\hstc\hstc\Frontend\python",'jpeg')

    customer = CustomerSheet(
                    s_no=row[0],
                    list_no=row[1],
                    sublist_no=row[2],
                    desc=row[10],
                    model=row[11],
                    dimensions=row[13],
                    remarks=row[14],
                    unit=row[15],
                    unit_price=row[16] ,
                    qty=row[17],
                    total_amt=row[18], 
                    discount_amt=row[19] ,
                    token_cus=row[25] ,
                    token_hstc=row[26] ,
                    token_date=row[27] ,
                    cbm=row[38] ,
                    ctns=row[39] ,
                    gross_wt=row[40] ,
                    net_wt=row[41],
                    photos=row[12] 	)
    customer_sheet.append(customer)

print(customer_sheet)
